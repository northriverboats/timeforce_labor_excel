#!/usr/bin/env python

from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font


class ExcelOpenDocument(object):
    """
    Some convenience methods for Excel OpenPyXL Documents
    """
    filename = None
    filename_saveas = None
    workbook = None
    sheet = None

    def __init__(self):
        pass

    def new(self, filename=None):
        self.filename_saveas = filename
        self.workbook = Workbook()
        self.sheet = self.workbook.active

    def open(self, filename=None):
        assert (filename), "No filename provided to open spreadsheet"
        self.filename = filename
        self.workbook = load_workbook(filename)
        self.sheet = self.workbook.active

    def from_template(self, filename=None, filename_saveas=None):
        assert (filename), "No filename provided to open as a template"
        assert (filename_saveas), "No filename provided to save as new spreadsheet"  # noqa: E501
        self.filename = filename_saveas
        self.workbook = load_workbook(filename)
        self.sheet = self.workbook.active

    def close(self):
        self.workbook = None
        self.sheet = None
        self.filename = None
        self.filename_saveas = None

    def saveas(self, filename):
        self.filename = filename
        self.filename_saveas = None
        self.workbook.save(filename)

    def save(self):
        assert (self.filename or self.filename_saveas), "Excel Document can not be saved: no filename given"  # noqa: E501
        if self.filename is not None:
            print("Saving '" + self.filename + "'")
            self.saveas(self.filename)
        else:
            print("Saving As '" + self.filename_saveas + "'")
            self.saveas(self.filename_saveas)

    def max_row(self):
        return self.sheet.max_row

    def max_column(self):
        return self.sheet.max_column

    def cell(self, ref=None, row=None, column=None):
        if ref:
            assert(not row or not column), 'Can not use column= or row= with "A1" style cell refernce'  # noqa: E501
            return self.sheet[ref]
        else:
            assert(not ref), 'Can not use "A1" style cell refernce with column= row='  # noqa: E501
            return self.sheet.cell(row=row, column=column)

    def font(self, **params):
        return Font(**params)

    def get_width(self, column):
        return self.sheet.column_dimensions[column].width

    def set_width(self, column, width):
        self.sheet.column_dimensions[column].width = width

